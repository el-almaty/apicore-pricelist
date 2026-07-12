"""
Скрипт собирает прайс-листы от нескольких дистрибьюторов apicore.kz:
товары + закупочные цены + остатки -> один Excel-файл, каждый
дистрибьютор на отдельном листе, и заливает результат на Диск Bitrix24.

Все секреты (ключи, ID) берутся из переменных окружения:
- APICORE_API_KEY
- APICORE_CATALOG_CODE   (необязательно, по умолчанию "main")
- APICORE_DISTRIBUTORS   -- JSON-список дистрибьюторов, например:
    [{"id": "3abb4152", "name": "Дистрибьютор А"},
     {"id": "b4d0d204", "name": "Дистрибьютор Б"}]
- BITRIX_WEBHOOK
- BITRIX_FOLDER_ID

Для GitHub Actions эти переменные передаются через Secrets репозитория
(см. .github/workflows/update-pricelist.yml).
"""

import os
import re
import json
import base64
import requests
import pandas as pd
import time
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule

API_KEY = os.environ["APICORE_API_KEY"]
CATALOG_CODE = os.environ.get("APICORE_CATALOG_CODE", "main")
DISTRIBUTORS = json.loads(os.environ["APICORE_DISTRIBUTORS"])

BASE_URL = "https://api.apicore.one/dealer"
HEADERS = {
    "Api-Key": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# --- Bitrix24 ---
BITRIX_WEBHOOK = os.environ["BITRIX_WEBHOOK"]  # напр. https://xxx.bitrix24.kz/rest/xx/xxxxxxxx/
BITRIX_FOLDER_ID = os.environ["BITRIX_FOLDER_ID"]


ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def clean_str(value):
    """Убирает управляющие символы, которые Excel-формат (.xlsx) физически
    не может хранить в ячейке -- иначе openpyxl падает с IllegalCharacterError."""
    if isinstance(value, str):
        return ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def call(version, method, body, max_retries=3):
    """Делает POST-запрос к apicore и возвращает JSON-ответ.
    При сбое (таймаут, временная перегрузка сервера и т.п.) повторяет попытку
    с нарастающей паузой, прежде чем сдаться."""
    url = f"{BASE_URL}/{version}/{method}"
    last_exc = None
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(url, headers=HEADERS, json=body, timeout=60)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            last_exc = e
            wait = attempt * 5  # 5, 10, 15 секунд
            print(f"    Попытка {attempt}/{max_retries} для {method} не удалась ({e}), жду {wait} сек...")
            time.sleep(wait)
    raise last_exc


def category_chain(cat_id, categories_by_id):
    """Возвращает список названий категорий от корня до текущей (полный путь).
    Идёт вверх по parent_id, пока не дойдёт до корня (parent_id == 0).
    Для товара без категории возвращает ["Без категории"]."""
    if cat_id not in categories_by_id or cat_id == 0:
        return ["Без категории"]

    chain = []
    node = categories_by_id.get(cat_id)
    seen = set()
    while node is not None and node["id"] not in seen:
        seen.add(node["id"])
        chain.append(clean_str(node["name"]))
        parent_id = node.get("parent_id", 0)
        if not parent_id:
            break
        node = categories_by_id.get(parent_id)

    chain.reverse()  # теперь от корня к текущей категории
    return chain


def fetch_all_pages(version, method, body, items_key):
    """Собирает ВСЕ элементы метода, даже если ответ приходит по частям
    (постранично, через offset/next_offset/total). Для небольших дистрибьюторов,
    у которых всё влезает в один ответ, отработает как обычный одиночный запрос."""
    all_items = []
    offset = 0
    while True:
        page_body = dict(body)
        page_body["offset"] = offset
        resp = call(version, method, page_body)
        items = resp.get(items_key, [])
        all_items.extend(items)

        total = resp.get("total")
        next_offset = resp.get("next_offset")
        count = resp.get("count", len(items))

        if total is not None and total > len(items) and (len(all_items) < total or next_offset is not None):
            print(f"    страница: получено {len(all_items)} из {total}")

        # условия остановки: нет информации о постраничности, следующей
        # страницы нет, страница пустая, или offset не продвигается (защита от зацикливания)
        if total is None or next_offset is None or not items or next_offset <= offset:
            break
        offset = next_offset
        time.sleep(0.7)  # лимит apicore -- не больше 2 запросов в секунду

    return all_items


def fetch_distributor_df(distributor_id, run_started_at):
    """Собирает полный прайс-лист (категории + товары + цены + остатки)
    для одного дистрибьютора и возвращает готовый DataFrame."""
    body = {"distributor_id": distributor_id, "catalog_code": CATALOG_CODE}

    print("  Получаю список категорий...")
    categories = fetch_all_pages("v1", "distrib.category.list", body, "categories")
    categories_by_id = {c["id"]: c for c in categories}
    print(f"    категорий получено: {len(categories_by_id)}")
    time.sleep(0.7)

    print("  Получаю список товаров...")
    products = fetch_all_pages("v1", "distrib.product.list", body, "products")
    print(f"    товаров получено: {len(products)}")
    time.sleep(0.7)

    print("  Получаю цены...")
    prices_list = fetch_all_pages("v2", "distrib.product.prices", body, "products")
    prices_by_id = {}
    price_updated_by_id = {}
    for p in prices_list:
        prices_by_id[p["product_id"]] = p["purchase"]["price"]
        if "date_update" in p:
            price_updated_by_id[p["product_id"]] = p["date_update"]
    time.sleep(0.7)

    print("  Получаю остатки...")
    qty_list = fetch_all_pages("v1", "distrib.product.quantities", body, "products")
    qty_by_id = {}
    qty_updated_by_id = {}
    for p in qty_list:
        qty_by_id[p["product_id"]] = p["quantity"]
        if "date_update" in p:
            qty_updated_by_id[p["product_id"]] = p["date_update"]
    time.sleep(0.7)

    # Сначала считаем цепочку категорий для каждого товара и находим
    # максимальную глубину вложенности именно у ЭТОГО дистрибьютора --
    # число колонок категорий подстраивается под неё, но не больше 3:
    # если уровней 4 и больше, всё, что глубже 2-го, склеивается в 3-ю колонку через " > "
    MAX_CAT_COLS = 3
    chains_by_pid = {p["id"]: category_chain(p.get("category_id", 0), categories_by_id) for p in products}
    max_depth = max((len(c) for c in chains_by_pid.values()), default=1)
    col_count = max(min(max_depth, MAX_CAT_COLS), 1)

    if col_count == 1:
        cat_col_names = ["Категория"]
    else:
        cat_col_names = [f"Категория {i} ур." for i in range(1, col_count + 1)]

    def to_display_chain(chain):
        if len(chain) <= col_count:
            return chain + [""] * (col_count - len(chain))
        # уровней больше, чем колонок -- последняя колонка вбирает всё, что глубже
        return chain[: col_count - 1] + [" > ".join(chain[col_count - 1:])]

    rows = []
    for p in products:
        pid = p["id"]
        chain = chains_by_pid[pid]
        display_chain = to_display_chain(chain)
        row = {}
        for col_name, value in zip(cat_col_names, display_chain):
            row[col_name] = value
        row.update({
            "ID": pid,
            "Наименование": clean_str(p["name"]),
            "Производитель": clean_str(p.get("vendor")) or "-",
            "Артикул": clean_str(p.get("vendor_code", "")),
            "Цена, KZT": prices_by_id.get(pid),
            "Наличие, шт": qty_by_id.get(pid, 0),
            "Штрихкод": clean_str(p.get("barcode", "")),
            "NTIN": clean_str(p.get("ntin", "")),
            "Данные получены": run_started_at,
        })
        rows.append(row)

    df = pd.DataFrame(rows)
    sort_cols = cat_col_names + ["Наименование"]
    df = df.sort_values(by=sort_cols, key=lambda col: col.str.lower())

    # если у ЭТОГО дистрибьютора производитель не указан вообще ни у одного
    # товара -- колонка только мешает (только прочерки), убираем её
    if (df["Производитель"] == "-").all():
        df = df.drop(columns=["Производитель"])

    # то же самое для штрихкода и NTIN -- если пусто у всех товаров, колонка не нужна
    for col in ["Штрихкод", "NTIN"]:
        if (df[col].astype(str).str.strip() == "").all():
            df = df.drop(columns=[col])

    return df


def sanitize_sheet_name(name):
    """Названия листов Excel не могут содержать [ ] : * ? / \\ и длиннее 31 символа."""
    clean = re.sub(r"[\[\]:*?/\\]", "", name).strip()
    return clean[:31] if clean else "Лист"


def format_worksheet(ws, df):
    """Оформляет один лист как Excel-таблицу (Table): фильтры, жирная шапка,
    полосатая заливка, подходит как источник для сводных таблиц.
    Плюс подбирает ширину столбцов по содержимому и формат чисел."""
    n_rows = ws.max_row
    n_cols = ws.max_column
    last_col_letter = get_column_letter(n_cols)

    table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title) or "Table"
    table = Table(displayName=f"T_{table_name}"[:255], ref=f"A1:{last_col_letter}{n_rows}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    # По умолчанию ширина колонки = длина самого длинного значения (+запас),
    # но для колонок категорий задан свой потолок -- иначе они растягивают лист
    header_by_col_idx = {c: cell.value for c, cell in enumerate(ws[1], start=1)}

    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in ws[letter]
        )
        header = header_by_col_idx.get(col_idx) or ""
        if header.startswith("Категория"):
            cap = 35
        elif header == "Артикул":
            cap = 40
        else:
            cap = 60
        ws.column_dimensions[letter].width = min(max_len + 2, cap)

    price_col = df.columns.get_loc("Цена, KZT") + 1
    qty_col = df.columns.get_loc("Наличие, шт") + 1
    for row in range(2, n_rows + 1):
        ws.cell(row=row, column=price_col).number_format = "#,##0"
        ws.cell(row=row, column=qty_col).number_format = "#,##0"


def build_summary_sheet(wb, distributors_and_sheets, run_started_dt):
    """Создаёт первый лист 'Сводка': по каждому дистрибьютору -- когда получены
    данные и сколько часов назад, с цветовой индикацией:
    зелёный -- обновлено < 1 часа назад
    жёлтый  -- от 1 до 24 часов назад
    красный -- сутки и более назад (вероятно, автообновление сломалось)"""
    ws = wb.create_sheet("Сводка", 0)

    ws["A1"] = "Дистрибьютор"
    ws["B1"] = "Данные получены"
    ws["C1"] = "Обновлено, часов назад"
    for cell in ws[1]:
        cell.font = Font(bold=True)

    row = 2
    for name, sheet_name in distributors_and_sheets:
        ws.cell(row=row, column=1, value=name)
        cell_b = ws.cell(row=row, column=2, value=run_started_dt.replace(tzinfo=None))
        cell_b.number_format = "dd.mm.yyyy hh:mm:ss"
        cell_c = ws.cell(row=row, column=3, value=f"=NOW()-B{row}")
        cell_c.number_format = "[h]:mm"
        # кликабельная ссылка на лист этого дистрибьютора
        ws.cell(row=row, column=1).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row, column=1).font = Font(color="0563C1", underline="single")
        row += 1

    last_row = row - 1
    if last_row >= 2:
        rng = f"C2:C{last_row}"
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        # порядок важен: правила проверяются по очереди, stopIfTrue прерывает
        # проверку на первом совпавшем условии
        rule_red = CellIsRule(operator="greaterThan", formula=["1"], fill=red, stopIfTrue=True)
        rule_yellow = CellIsRule(operator="greaterThan", formula=["1/24"], fill=yellow, stopIfTrue=True)
        rule_green = CellIsRule(operator="lessThanOrEqual", formula=["1/24"], fill=green, stopIfTrue=True)
        ws.conditional_formatting.add(rng, rule_red)
        ws.conditional_formatting.add(rng, rule_yellow)
        ws.conditional_formatting.add(rng, rule_green)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.freeze_panes = "A2"


def find_existing_file_id(folder_id, filename):
    """Ищет в папке файл с указанным именем. Возвращает его ID или None, если не найден."""
    url = f"{BITRIX_WEBHOOK}disk.folder.getchildren"
    resp = requests.post(url, data={"id": folder_id}, timeout=60)
    resp.raise_for_status()
    result = resp.json()
    for item in result.get("result", []):
        if item.get("TYPE") == "file" and item.get("NAME") == filename:
            return item["ID"]
    return None


def upload_to_bitrix_disk(local_file_path, folder_id, filename):
    """Загружает файл в указанную папку на Диске Bitrix24.
    Если файл с таким именем уже существует -- обновляет его СОДЕРЖИМОЕ
    через disk.file.uploadversion (передаётся сразу base64, одним запросом),
    благодаря чему ID файла и ссылка на него не меняются от запуска к запуску.
    Если файла ещё нет -- создаёт его через disk.folder.uploadfile
    (двухшаговая схема с uploadUrl, только в самый первый раз)."""
    existing_file_id = find_existing_file_id(folder_id, filename)

    if existing_file_id:
        print(f"Файл '{filename}' уже существует (ID {existing_file_id}) -- обновляю версию.")
        with open(local_file_path, "rb") as f:
            content_b64 = base64.b64encode(f.read()).decode("ascii")

        data = {
            "id": existing_file_id,
            "fileContent[0]": filename,
            "fileContent[1]": content_b64,
        }
        resp = requests.post(f"{BITRIX_WEBHOOK}disk.file.uploadversion", data=data, timeout=120)
        if not resp.ok:
            print("Bitrix24 вернул ошибку, тело ответа:", resp.text)
        resp.raise_for_status()
        result = resp.json()
        print("Ответ Bitrix24:", result)
        if "result" not in result:
            raise RuntimeError(f"Bitrix24 вернул ошибку при обновлении версии: {result}")
        print("Файл на Диске Bitrix24 обновлён.")
        return result["result"]

    print(f"Файла '{filename}' ещё нет -- создаю новый.")
    step1_url = f"{BITRIX_WEBHOOK}disk.folder.uploadfile"
    data = {
        "id": folder_id,
        "data[NAME]": filename,
    }
    resp1 = requests.post(step1_url, data=data, timeout=60)
    if not resp1.ok:
        print("Bitrix24 (шаг 1) вернул ошибку, тело ответа:", resp1.text)
    resp1.raise_for_status()
    step1_result = resp1.json()
    print("Шаг 1 (получение uploadUrl):", step1_result)

    if "result" not in step1_result or "uploadUrl" not in step1_result["result"]:
        raise RuntimeError(f"Не удалось получить uploadUrl: {step1_result}")

    upload_url = step1_result["result"]["uploadUrl"]

    with open(local_file_path, "rb") as f:
        files = {"file": (filename, f)}
        resp2 = requests.post(upload_url, files=files, timeout=120)
    if not resp2.ok:
        print("Bitrix24 (шаг 2) вернул ошибку, тело ответа:", resp2.text)
    resp2.raise_for_status()
    step2_result = resp2.json()
    print("Шаг 2 (загрузка содержимого):", step2_result)

    if "result" not in step2_result:
        raise RuntimeError(f"Bitrix24 вернул ошибку при загрузке файла: {step2_result}")

    print("Файл загружен на Диск Bitrix24.")
    return step2_result["result"]


def main():
    run_started_dt = datetime.now(ZoneInfo("Asia/Almaty"))
    run_started_at = run_started_dt.strftime("%d.%m.%Y %H:%M:%S")
    out_file = "pricelist.xlsx"

    dfs_by_sheet = {}
    succeeded_names = []
    failed = []
    for d in DISTRIBUTORS:
        print(f"=== Дистрибьютор: {d['name']} ({d['id']}) ===")
        try:
            df = fetch_distributor_df(d["id"], run_started_at)
        except Exception as e:
            print(f"  ОШИБКА при обработке {d['name']} ({d['id']}): {e}")
            failed.append(d["name"])
            continue
        sheet_name = sanitize_sheet_name(d["name"])
        # на случай совпадения названий после обрезки/очистки -- не даём листам дублироваться
        base_name, i = sheet_name, 2
        while sheet_name in dfs_by_sheet:
            suffix = f" ({i})"
            sheet_name = base_name[: 31 - len(suffix)] + suffix
            i += 1
        dfs_by_sheet[sheet_name] = df
        succeeded_names.append((d["name"], sheet_name))
        print(f"  готово: {len(df)} строк")

    if not dfs_by_sheet:
        raise RuntimeError("Ни один дистрибьютор не обработан успешно -- файл не создан.")

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        for sheet_name, df in dfs_by_sheet.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

    wb = load_workbook(out_file)
    for sheet_name, df in dfs_by_sheet.items():
        format_worksheet(wb[sheet_name], df)
    build_summary_sheet(wb, succeeded_names, run_started_dt)
    wb.save(out_file)

    total_rows = sum(len(df) for df in dfs_by_sheet.values())
    print(f"Готово. Файл сохранён: {out_file} ({len(dfs_by_sheet)} листов, {total_rows} строк всего)")

    if BITRIX_FOLDER_ID:
        upload_to_bitrix_disk(out_file, BITRIX_FOLDER_ID, "Прайс-лист.xlsx")
    else:
        print("BITRIX_FOLDER_ID не задан -- пропускаю загрузку на Bitrix24 Диск.")

    if failed:
        print(f"ВНИМАНИЕ: не удалось обработать {len(failed)} дистрибьютор(ов): {', '.join(failed)}")
        raise SystemExit(1)  # файл всё равно загружен, но запуск помечается как проблемный


if __name__ == "__main__":
    main()
